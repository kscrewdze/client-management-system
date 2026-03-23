# -*- coding: utf-8 -*-

"""Экспорт в Excel с сохранением форматирования как в программе"""
import os
from datetime import datetime
from gui.frames.export.exporters.base_exporter import BaseExporter


class ExcelExporter(BaseExporter):
    """Экспортер в Excel с форматированием как в программе"""
    
    def export(self):
        """Экспорт в Excel"""
        try:
            # Проверка наличия библиотеки openpyxl
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                self.notifier.show_error("❌ Библиотека openpyxl не установлена")
                return
            
            # Проверка данных
            clients = self.check_data()
            if not clients:
                return
            
            # Создание файла
            filename = self.get_export_filename("xlsx")
            
            # Создание рабочей книги
            wb = Workbook()
            ws = wb.active
            ws.title = "Клиенты"
            
            # Заголовки как в программе
            headers = [
                "ID", "Имя", "Telegram", "Телефон", "Дата рождения",
                "Число судьбы", "Матрица", "Цена", "Дата заказа",
                "Статус", "Дата выполнения", "Комментарий"
            ]
            
            # Стили для заголовков (как в теме)
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2b5e8c", end_color="2b5e8c", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Заполняем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            # Стили для данных
            data_font = Font(size=10)
            data_alignment = Alignment(horizontal="left", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")
            price_alignment = Alignment(horizontal="right", vertical="center")
            
            completed_fill = PatternFill(start_color="e8f5e8", end_color="e8f5e8", fill_type="solid")
            
            # Данные
            for row, client in enumerate(clients, 2):
                # ID (центрированный)
                cell = ws.cell(row=row, column=1, value=client.id)
                cell.font = data_font
                cell.alignment = center_alignment
                
                # Имя
                cell = ws.cell(row=row, column=2, value=client.name)
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Telegram
                cell = ws.cell(row=row, column=3, value=client.telegram or "")
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Телефон
                cell = ws.cell(row=row, column=4, value=client.phone or "")
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Дата рождения
                cell = ws.cell(row=row, column=5, value=client.birth_date)
                cell.font = data_font
                cell.alignment = center_alignment
                
                # Число судьбы
                cell = ws.cell(row=row, column=6, value=client.destiny_number)
                cell.font = data_font
                cell.alignment = center_alignment
                
                # Матрица
                cell = ws.cell(row=row, column=7, value=client.matrix_name or "")
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Цена (с форматированием)
                price_text = f"{client.service_price:,.0f} ₽".replace(",", " ")
                cell = ws.cell(row=row, column=8, value=price_text)
                cell.font = data_font
                cell.alignment = price_alignment
                
                # Дата заказа
                cell = ws.cell(row=row, column=9, value=client.order_date)
                cell.font = data_font
                cell.alignment = center_alignment
                
                # Статус
                status_text = "✅ Выполнен" if client.is_completed else "⬜ В ожидании"
                cell = ws.cell(row=row, column=10, value=status_text)
                cell.font = data_font
                cell.alignment = center_alignment
                
                # Дата выполнения
                completed_date = client.completed_date or ""
                if client.is_completed and completed_date:
                    completed_text = f"✅ {completed_date}"
                else:
                    completed_text = ""
                cell = ws.cell(row=row, column=11, value=completed_text)
                cell.font = data_font
                cell.alignment = center_alignment
                
                # Комментарий
                cell = ws.cell(row=row, column=12, value=client.comment or "")
                cell.font = data_font
                cell.alignment = data_alignment
                
                # Применяем границы и цвет для выполненных заказов
                for col in range(1, 13):
                    cell = ws.cell(row=row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if client.is_completed:
                        cell.fill = completed_fill
            
            # Автоподбор ширины колонок как в программе
            column_widths = {
                1: 8,    # ID
                2: 25,   # Имя
                3: 20,   # Telegram
                4: 18,   # Телефон
                5: 13,   # Дата рождения
                6: 10,   # Число судьбы
                7: 25,   # Матрица
                8: 15,   # Цена
                9: 13,   # Дата заказа
                10: 15,  # Статус
                11: 15,  # Дата выполнения
                12: 30   # Комментарий
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Заморозка первой строки (заголовки)
            ws.freeze_panes = 'A2'
            
            # Добавляем фильтры
            ws.auto_filter.ref = f"A1:L{len(clients) + 1}"
            
            # Сохранение
            wb.save(str(filename))
            
            self.notifier.show_success(f"✅ Экспортировано {len(clients)} записей в Excel")
            
            # Спрашиваем, открыть ли папку
            if self.ask_open_folder(filename):
                self.open_export_folder()
            
            return filename
            
        except Exception as e:
            self.notifier.show_error(f"❌ Ошибка экспорта в Excel: {str(e)}")
            return None