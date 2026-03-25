# -*- coding: utf-8 -*-

"""Автоэкспорт данных в Excel (clients.xlsx / matrices.xlsx)."""
import logging
import os
from pathlib import Path
from typing import Any, List

from config.settings import Settings
from database.models import Client, Matrix

logger = logging.getLogger(__name__)

_DIR = Settings.EXPORTS_DIR


def _ensure_dir() -> None:
    _DIR.mkdir(exist_ok=True)


def export_clients(db: Any) -> Path:
    """Экспорт всех клиентов в exports/clients.xlsx (перезапись)."""
    _ensure_dir()
    path = _DIR / "clients.xlsx"
    clients = db.get_all_clients() or []

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl не установлен")
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = [
        "ID", "Имя", "Telegram", "Телефон", "Дата рожд.",
        "ЧС", "Матрица", "Цена", "Дата заказа", "Статус", "Комментарий",
    ]
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2B5E8C", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = center
        cell.border = border

    for r, c in enumerate(clients, 2):
        vals = [
            c.id, c.name, c.telegram or "", c.phone or "",
            c.birth_date, c.destiny_number, c.matrix_name or "",
            c.service_price, c.order_date,
            "Выполнен" if c.is_completed else "В ожидании",
            c.comment or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(length, 40)

    wb.save(str(path))
    logger.info("Авто-экспорт клиентов → %s (%d)", path.name, len(clients))
    return path


def export_matrices(db: Any) -> Path:
    """Экспорт всех матриц в exports/matrices.xlsx (перезапись)."""
    _ensure_dir()
    path = _DIR / "matrices.xlsx"
    matrices = db.get_all_matrices() or []

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl не установлен")
        return path

    wb = Workbook()
    ws = wb.active
    ws.title = "Матрицы"

    headers = ["ID", "Название", "Цена", "Дата создания"]
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2e7d32", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = center
        cell.border = border

    for r, m in enumerate(matrices, 2):
        vals = [m.id, m.name, m.price, (m.created_date or "")[:10]]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(length, 40)

    wb.save(str(path))
    logger.info("Авто-экспорт матриц → %s (%d)", path.name, len(matrices))
    return path


def auto_export_all(db: Any) -> None:
    """Экспорт и клиентов, и матриц (вызывается после любого изменения)."""
    try:
        export_clients(db)
        export_matrices(db)
    except Exception as e:
        logger.error("Ошибка авто-экспорта: %s", e)
